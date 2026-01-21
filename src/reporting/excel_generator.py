"""
Generador de informes Excel con estadísticas.
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference

import sys
sys.path.insert(0, str(__file__).rsplit("/", 4)[0])
from config.settings import OUTPUT_DIR
from src.extraction.models import Expediente
from src.reporting.csv_generator import expedientes_to_dataframe

logger = logging.getLogger(__name__)


def generate_excel_report(
    expedientes: list[Expediente],
    output_path: Optional[Path] = None,
    filename: str = "informe_cnmc.xlsx",
) -> Path:
    """
    Genera un informe Excel completo con datos y estadísticas.

    Args:
        expedientes: Lista de expedientes
        output_path: Directorio de salida
        filename: Nombre del archivo

    Returns:
        Path del archivo generado
    """
    output_dir = output_path or OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    filepath = output_dir / filename

    # Crear workbook
    wb = Workbook()

    # Hoja 1: Datos detallados
    ws_data = wb.active
    ws_data.title = "Expedientes"

    df = expedientes_to_dataframe(expedientes)

    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)

            # Formato para cabecera
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

    # Ajustar anchos de columna
    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[column_letter].width = adjusted_width

    # Hoja 2: Resumen estadístico
    ws_stats = wb.create_sheet("Estadísticas")

    # Calcular estadísticas
    total = len(df)
    stats = df["Resultado_clasificado"].value_counts()

    # Título
    ws_stats["A1"] = "RESUMEN DE CLASIFICACIÓN"
    ws_stats["A1"].font = Font(bold=True, size=14)
    ws_stats.merge_cells("A1:C1")

    # Tabla de estadísticas
    ws_stats["A3"] = "Resultado"
    ws_stats["B3"] = "Cantidad"
    ws_stats["C3"] = "Porcentaje"

    for cell in ["A3", "B3", "C3"]:
        ws_stats[cell].font = Font(bold=True)
        ws_stats[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_stats[cell].font = Font(bold=True, color="FFFFFF")

    row = 4
    for resultado, count in stats.items():
        percentage = (count / total * 100) if total > 0 else 0
        ws_stats[f"A{row}"] = resultado
        ws_stats[f"B{row}"] = count
        ws_stats[f"C{row}"] = f"{percentage:.1f}%"
        row += 1

    # Total
    ws_stats[f"A{row}"] = "TOTAL"
    ws_stats[f"B{row}"] = total
    ws_stats[f"C{row}"] = "100%"
    ws_stats[f"A{row}"].font = Font(bold=True)
    ws_stats[f"B{row}"].font = Font(bold=True)
    ws_stats[f"C{row}"].font = Font(bold=True)

    # Gráfico de barras
    if len(stats) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Distribución de Resultados"
        chart.y_axis.title = "Cantidad"
        chart.x_axis.title = "Resultado"

        data = Reference(ws_stats, min_col=2, min_row=3, max_row=row - 1, max_col=2)
        cats = Reference(ws_stats, min_col=1, min_row=4, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 15
        chart.height = 10

        ws_stats.add_chart(chart, "E3")

    # Gráfico circular
    if len(stats) > 0:
        pie = PieChart()
        pie.title = "Porcentaje por Categoría"

        data = Reference(ws_stats, min_col=2, min_row=3, max_row=row - 1, max_col=2)
        labels = Reference(ws_stats, min_col=1, min_row=4, max_row=row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 12
        pie.height = 10

        ws_stats.add_chart(pie, "E18")

    # Ajustar anchos
    ws_stats.column_dimensions["A"].width = 20
    ws_stats.column_dimensions["B"].width = 12
    ws_stats.column_dimensions["C"].width = 12

    # Guardar
    wb.save(filepath)
    logger.info(f"Excel generado: {filepath}")

    return filepath
